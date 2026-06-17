#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
claim_export.py — for a given month, build one Excel file per employee
(named Name_YYYY-MM.xlsx). Each workbook has 3 tabs: Taxi / Food / Other,
each listing that category's claims plus a subtotal.

Receipt photos are compressed and uploaded to the Cloudflare receipt-hosting
Worker; the Excel only stores a clickable "📎 View receipt" link (keeping the
sheet clean; links expire after ~60 days by default, while the originals stay
in receipts/ forever). After generation, all employee files are pushed to the
finance WhatsApp via the proactive queue (drained by the bridge).

Usage:
    python3 claim_export.py                  # previous month, all employees
    python3 claim_export.py 2026-06          # specific month, all employees
    python3 claim_export.py 2026-06 Monica   # specific month, only Monica
"""

import os
import io
import re
import sys
import json
import base64
import sqlite3
import urllib.request
from collections import OrderedDict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "claim_config.json")

# (db type, display title, sheet name)
CATEGORIES = [
    ("taxi", "🚕 Taxi", "Taxi"),
    ("food", "🍱 Food", "Food"),
    ("other", "📦 Other", "Other"),
]
TYPE_PLAIN = {"taxi": "Taxi", "food": "Food", "other": "Other"}
LINK_FONT = Font(color="0563C1", underline="single")


def load_config():
    cfg = {
        "finance_whatsapp": "",
        "currency": "USD",
        "db_path": "claims.db",
        "proactive_queue_path": "/tmp/wa_proactive_queue.json",
        "receipt_host_url": "",
        "receipt_upload_secret": "",
        "receipt_ttl_days": 60,
    }
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            cfg.update(json.load(f))
    except Exception as e:
        print(f"[Export] Failed to read claim_config.json: {e}")
    return cfg


def prev_month_period(now=None):
    now = now or datetime.now()
    y, m = now.year, now.month
    return f"{y-1}-12" if m == 1 else f"{y}-{m-1:02d}"


def fetch_claims(db_path, period):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM claims WHERE substr(claim_date,1,7) = ? "
            "ORDER BY employee_name, claim_date",
            (period,),
        ).fetchall()
    finally:
        conn.close()
    return rows


# -- Receipts: compress + upload to the Worker (with local cache) ---------
def _ensure_upload_table(conn):
    conn.execute(
        """CREATE TABLE IF NOT EXISTS receipt_uploads (
               receipt_path TEXT PRIMARY KEY,
               url TEXT NOT NULL,
               uploaded_at TEXT NOT NULL
           )"""
    )


def _compress_image(path):
    from PIL import Image
    im = Image.open(path)
    if im.mode != "RGB":
        im = im.convert("RGB")
    im.thumbnail((1280, 1280))
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=72, optimize=True)
    return buf.getvalue()


def _http_upload(host, secret, ttl_days, data, meta):
    meta_b64 = base64.b64encode(
        json.dumps(meta, ensure_ascii=False).encode("utf-8")
    ).decode("ascii")
    req = urllib.request.Request(
        host.rstrip("/") + "/upload",
        data=data,
        method="POST",
        headers={
            "X-Auth": secret,
            "Content-Type": "image/jpeg",
            "X-TTL-Days": str(ttl_days),
            "X-Meta": meta_b64,
            # A normal UA avoids Cloudflare's bot block (error 1010)
            "User-Agent": "claim-bot/1.0",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode()).get("url")


def upload_receipt(db_path, path, meta, cfg):
    """Return the receipt's view link; returns None if unavailable
    (no error — the report is still generated)."""
    host = (cfg.get("receipt_host_url") or "").strip()
    secret = (cfg.get("receipt_upload_secret") or "").strip()
    ttl_days = int(cfg.get("receipt_ttl_days", 60))
    if not host or not secret or not path or not os.path.exists(path):
        return None

    conn = sqlite3.connect(db_path)
    try:
        _ensure_upload_table(conn)
        row = conn.execute(
            "SELECT url, uploaded_at FROM receipt_uploads WHERE receipt_path = ?",
            (path,),
        ).fetchone()
        if row:
            try:
                age_days = (datetime.now() - datetime.fromisoformat(row[1])).days
                if age_days < max(1, ttl_days - 7):  # still valid (7-day margin)
                    return row[0]
            except Exception:
                pass
        try:
            data = _compress_image(path)
            url = _http_upload(host, secret, ttl_days, data, meta)
        except Exception as e:
            print(f"[Export] Failed to upload receipt ({path}): {e}")
            return None
        if url:
            conn.execute(
                "INSERT OR REPLACE INTO receipt_uploads (receipt_path, url, uploaded_at) "
                "VALUES (?,?,?)",
                (path, url, datetime.now().isoformat()),
            )
            conn.commit()
        return url
    finally:
        conn.close()


def build_employee_workbook(name, period, emp_rows, currency, out_path, url_map):
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we create our own 3

    by_type = {"taxi": [], "food": [], "other": []}
    for r in emp_rows:
        by_type.get(r["type"], by_type["other"]).append(r)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Date", f"Amount ({currency})", "Note", "Status", "Receipt"]
    widths = [13, 16, 34, 12, 18]

    for key, label, sheet_name in CATEGORIES:
        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value=f"{name} — {period} — {label}").font = Font(bold=True, size=13)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        r = 3
        subtotal = 0.0
        for row in by_type[key]:
            ws.cell(row=r, column=1, value=row["claim_date"])
            amt = ws.cell(row=r, column=2, value=round(row["amount"], 2))
            amt.number_format = "#,##0.00"
            ws.cell(row=r, column=3, value=row["note"] or "")
            ws.cell(row=r, column=4, value=row["status"])

            rp = row["receipt_path"]
            url = url_map.get(rp) if rp else None
            cell = ws.cell(row=r, column=5)
            if url:
                cell.value = "📎 View receipt"
                cell.hyperlink = url
                cell.font = LINK_FONT
            elif rp:
                cell.value = "(receipt on file)"
            else:
                cell.value = "(no receipt)"

            subtotal += row["amount"]
            r += 1

        ws.cell(row=r + 1, column=1, value="Subtotal").font = Font(bold=True)
        st = ws.cell(row=r + 1, column=2, value=round(subtotal, 2))
        st.font = Font(bold=True)
        st.number_format = "#,##0.00"
        ws.freeze_panes = "A3"

    wb.save(out_path)


def normalize_wa_id(value):
    """Bare number (e.g. 6591234567) / @c.us / @g.us -> a WhatsApp chat id."""
    v = (value or "").strip()
    if v.endswith("@c.us") or v.endswith("@g.us"):
        return v
    digits = re.sub(r"[^0-9]", "", v)
    return f"{digits}@c.us" if digits else ""


def send_whatsapp(queue_path, target_id, message, file_paths):
    """Write the report into the proactive queue; one message, multiple files.
    The WhatsApp bridge drains this queue and actually sends them."""
    payload = {
        "group_id": target_id,
        "message": message,
        "files": [
            {"path": p, "caption": os.path.basename(p), "type": "document"}
            for p in file_paths
        ],
    }
    existing = []
    try:
        with open(queue_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            existing = data if isinstance(data, list) else [data]
    except Exception:
        existing = []
    existing.append(payload)
    tmp = queue_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False)
    os.replace(tmp, queue_path)


def generate_and_send(period=None, employee_filter=None):
    """Build per-employee Excel files for a month (YYYY-MM, default previous)
    and queue them for sending. If employee_filter is set, only employees whose
    name matches (case-insensitive substring) are generated.
    Returns (ok, message)."""
    cfg = load_config()
    period = period or prev_month_period()
    db_path = os.path.join(BASE_DIR, cfg["db_path"])
    currency = cfg.get("currency", "USD")

    if not os.path.exists(db_path):
        return False, f"Database not found: {db_path}"

    rows = fetch_claims(db_path, period)
    if not rows:
        return True, f"No claims for {period}, nothing generated"

    groups = OrderedDict()
    for r in rows:
        g = groups.setdefault(r["wa_id"], {"name": r["employee_name"], "rows": []})
        g["rows"].append(r)

    if employee_filter:
        key = employee_filter.strip().lower()
        groups = OrderedDict(
            (k, v) for k, v in groups.items() if key in v["name"].lower()
        )
        if not groups:
            return True, f"No claims matching '{employee_filter}' for {period}"

    # Upload receipts -> receipt_path -> url map (cached, each path uploaded once)
    url_map = {}
    for g in groups.values():
        for r in g["rows"]:
            rp = r["receipt_path"]
            if rp and rp not in url_map:
                meta = {
                    "employee": g["name"],
                    "type": TYPE_PLAIN.get(r["type"], r["type"]),
                    "amount": f"{r['amount']:.2f}",
                    "currency": currency,
                    "date": r["claim_date"],
                }
                url_map[rp] = upload_receipt(db_path, rp, meta, cfg)

    out_dir = os.path.join(BASE_DIR, "exports", period)
    os.makedirs(out_dir, exist_ok=True)

    files = []
    summary_lines = []
    grand_total = 0.0
    for g in groups.values():
        name = g["name"]
        emp_rows = g["rows"]
        emp_total = sum(r["amount"] for r in emp_rows)
        grand_total += emp_total
        safe = re.sub(r"[^0-9A-Za-z]+", "_", name).strip("_") or "employee"
        out_path = os.path.join(out_dir, f"{safe}_{period}.xlsx")
        build_employee_workbook(name, period, emp_rows, currency, out_path, url_map)
        files.append(out_path)
        summary_lines.append(f"• {name}: {currency} {emp_total:,.2f}")

    print(f"[Export] {period}: generated {len(files)} employee report(s) -> {out_dir}, total {currency} {grand_total:.2f}")

    target = normalize_wa_id(cfg.get("finance_whatsapp", ""))
    if not target:
        return True, f"Generated {len(files)} report(s) to {out_dir} (finance_whatsapp not set, nothing sent)"

    message = (
        f"📊 *Expense claims — {period}*\n"
        + "\n".join(summary_lines)
        + f"\n\nTotal: {currency} {grand_total:,.2f}\n"
        + f"({len(files)} employee file(s), 3 tabs each: Taxi / Food / Other)\n"
        + "Open with a spreadsheet app and tap '📎 View receipt' to see each receipt."
    )
    try:
        send_whatsapp(cfg["proactive_queue_path"], target, message, files)
    except Exception as e:
        return False, f"Failed to write to the WhatsApp queue: {e}"
    return True, f"Queued {len(files)} report(s) to {target} (total {currency} {grand_total:.2f})"


if __name__ == "__main__":
    period_arg = sys.argv[1] if len(sys.argv) > 1 else None
    name_arg = sys.argv[2] if len(sys.argv) > 2 else None
    ok, message = generate_and_send(period_arg, name_arg)
    print(("✅ " if ok else "❌ ") + message)
    sys.exit(0 if ok else 1)
